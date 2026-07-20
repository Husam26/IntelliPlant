"""
Document Processing Pipeline — The AI core of IntelliPlant.

This module handles the entire document lifecycle:
1. TEXT EXTRACTION: Read text from PDF/DOCX/XLSX/TXT files
2. CHUNKING: Split long text into overlapping chunks (~500 chars each)
3. EMBEDDING: Convert chunks to vector embeddings (all-MiniLM-L6-v2)
4. STORAGE: Store embeddings in ChromaDB for similarity search
5. CLASSIFICATION: Auto-classify document type using keyword heuristics

WHY CHUNKING?
LLMs have context limits. We can't feed a 100-page manual to the AI.
Instead, we split it into small chunks and use vector similarity to find
the most relevant chunks for any given question. This is the "R" in RAG.

WHY EMBEDDINGS?
Embeddings convert text to numerical vectors (arrays of numbers).
Similar text produces similar vectors. So "pump seal failure" and
"mechanical seal leaked" will have similar embeddings, even though
the exact words differ. This enables semantic search.
"""

import os
import traceback
from typing import List, Optional

from app.config import settings
from app.database import SessionLocal
from app.models.models import Document


# ========================
# Text Extraction
# ========================

def extract_text_from_pdf(file_path: str) -> tuple[str, int]:
    """
    Extract text from a PDF file using PyPDF2.
    
    Returns: (extracted_text, page_count)
    
    PyPDF2 reads each page's text content. For scanned PDFs (images),
    this won't work — you'd need OCR (Tesseract). For the hackathon,
    we'll use text-based PDFs.
    """
    from PyPDF2 import PdfReader

    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            text += page_text + "\n\n"
    return text.strip(), len(reader.pages)


def extract_text_from_docx(file_path: str) -> tuple[str, int]:
    """
    Extract text from a Word document using python-docx.
    
    Reads all paragraphs and joins them. Page count is estimated
    from character count (rough: ~3000 chars per page).
    """
    from docx import Document as DocxDocument

    doc = DocxDocument(file_path)
    text = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
    page_count = max(1, len(text) // 3000)
    return text.strip(), page_count


def extract_text_from_xlsx(file_path: str) -> tuple[str, int]:
    """
    Extract text from an Excel file using openpyxl.
    
    Reads all sheets, all rows, all cells. Joins everything as text.
    Useful for maintenance logs and inspection checklists in spreadsheet format.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    text_parts = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        text_parts.append(f"=== Sheet: {sheet} ===")
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join([str(cell) for cell in row if cell is not None])
            if row_text.strip():
                text_parts.append(row_text)

    wb.close()
    text = "\n".join(text_parts)
    return text.strip(), len(wb.sheetnames)


def extract_text_from_txt(file_path: str) -> tuple[str, int]:
    """Extract text from a plain text file."""
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()
    page_count = max(1, len(text) // 3000)
    return text.strip(), page_count


def extract_text(file_path: str, file_type: str) -> tuple[str, int]:
    """
    Route to the correct extractor based on file type.
    This is the main entry point for text extraction.
    """
    extractors = {
        "pdf": extract_text_from_pdf,
        "docx": extract_text_from_docx,
        "doc": extract_text_from_docx,
        "xlsx": extract_text_from_xlsx,
        "xls": extract_text_from_xlsx,
        "txt": extract_text_from_txt,
        "csv": extract_text_from_txt,
    }

    extractor = extractors.get(file_type)
    if not extractor:
        raise ValueError(f"Unsupported file type: {file_type}")

    return extractor(file_path)


# ========================
# Text Chunking
# ========================

def chunk_text(text: str, chunk_size: int = 500, overlap: int = 100) -> List[str]:
    """
    Split text into overlapping chunks for embedding.
    
    WHY OVERLAP?
    If a key piece of information spans two chunks, overlap ensures
    it appears fully in at least one chunk. Without overlap, answers
    could be cut in half.
    
    Args:
        text: The full document text
        chunk_size: Characters per chunk (~500 = ~100 tokens)
        overlap: Characters of overlap between consecutive chunks
    
    Returns:
        List of text chunks
    """
    if not text:
        return []

    chunks = []
    start = 0

    while start < len(text):
        end = start + chunk_size

        # Try to break at a sentence boundary (period, newline)
        if end < len(text):
            # Look for the last period or newline in the chunk
            last_period = text.rfind(".", start, end)
            last_newline = text.rfind("\n", start, end)
            break_point = max(last_period, last_newline)

            if break_point > start + chunk_size // 2:
                end = break_point + 1  # Include the period

        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)

        start = end - overlap  # Move back by overlap amount

    return chunks


# ========================
# FAISS Vector Storage
# ========================

def get_faiss_index():
    """
    Get or create the FAISS index for document chunks.
    
    FAISS stores vectors (embeddings) alongside metadata (via LangChain).
    It handles the similarity search — given a query embedding,
    it finds the most similar document chunks.
    """
    import os
    from langchain_community.vectorstores import FAISS
    from langchain_community.embeddings import HuggingFaceEmbeddings
    
    embeddings = HuggingFaceEmbeddings(model_name=settings.EMBEDDING_MODEL)
    faiss_path = settings.FAISS_PERSIST_DIR
    
    if os.path.exists(faiss_path) and os.path.exists(os.path.join(faiss_path, "index.faiss")):
        return FAISS.load_local(faiss_path, embeddings, allow_dangerous_deserialization=True)
    else:
        # Create an empty index (we need at least one text to initialize, or use from_texts)
        # We will initialize it properly on first insert.
        return None


def store_chunks_in_faiss(doc_id: str, chunks: List[str], filename: str):
    """
    Store document chunks in FAISS with metadata.
    """
    if not chunks:
        return 0

    import os
    from langchain_community.vectorstores import FAISS
    from langchain_community.embeddings import HuggingFaceEmbeddings

    embeddings = HuggingFaceEmbeddings(model_name=settings.EMBEDDING_MODEL)
    faiss_path = settings.FAISS_PERSIST_DIR

    metadatas = [
        {
            "document_id": doc_id,
            "filename": filename,
            "chunk_index": i,
            "total_chunks": len(chunks),
        }
        for i in range(len(chunks))
    ]
    ids = [f"{doc_id}_chunk_{i}" for i in range(len(chunks))]

    vectorstore = get_faiss_index()
    if vectorstore is None:
        vectorstore = FAISS.from_texts(texts=chunks, embedding=embeddings, metadatas=metadatas, ids=ids)
    else:
        vectorstore.add_texts(texts=chunks, metadatas=metadatas, ids=ids)
        
    vectorstore.save_local(faiss_path)

    return len(chunks)


def delete_document_chunks(doc_id: str):
    """Remove chunks is slightly complex in raw FAISS, skipping for MVP."""
    pass


# ========================
# Document Classification
# ========================

def classify_document(text: str, filename: str) -> str:
    """
    Auto-classify document type using LLM.
    
    Categories:
    - SOP: Standard Operating Procedure
    - Incident Report: Incident/accident report
    - Maintenance Report: Maintenance report/log
    - Compliance Report: Compliance/audit report
    - Equipment Manual: Equipment manual/specifications
    - Uncategorized: Anything else
    """
    from app.services.llm_service import get_gemini_model
    from langchain_core.messages import HumanMessage, SystemMessage
    
    llm = get_gemini_model()
    if not llm:
        # Fallback to simple heuristics
        text_lower = (text + " " + filename).lower()
        if "procedure" in text_lower or "sop" in text_lower: return "SOP"
        if "incident" in text_lower or "injury" in text_lower: return "Incident Report"
        if "maintenance" in text_lower or "repair" in text_lower: return "Maintenance Report"
        if "compliance" in text_lower or "audit" in text_lower: return "Compliance Report"
        if "manual" in text_lower or "specification" in text_lower: return "Equipment Manual"
        return "Uncategorized"
        
    system_prompt = (
        "You are an industrial document classifier. Categorize the following document into exactly one of these classes: "
        "'SOP', 'Incident Report', 'Maintenance Report', 'Compliance Report', 'Equipment Manual', 'Uncategorized'. "
        "Reply with just the category name and nothing else."
    )
    
    try:
        response = llm.invoke([
            SystemMessage(content=system_prompt),
            HumanMessage(content=f"Filename: {filename}\n\nContent snippet:\n{text[:2000]}")
        ])
        cat = response.content.strip()
        allowed = ["SOP", "Incident Report", "Maintenance Report", "Compliance Report", "Equipment Manual", "Uncategorized"]
        if cat in allowed:
            return cat
        return "Uncategorized"
    except Exception as e:
        print(f"[ERROR] LLM Classification failed: {e}")
        return "Uncategorized"


# ========================
# Main Processing Pipeline
# ========================

def process_document(doc_id: str):
    """
    Main document processing pipeline — called as a background task.
    
    Flow:
    1. Load document record from DB
    2. Extract text from the file
    3. Classify the document type
    4. Chunk the text
    5. Store chunks + embeddings in ChromaDB
    6. Update the document record with results
    
    This runs in the background after upload, so the user gets an
    immediate response and can see the processing status update.
    """
    db = SessionLocal()

    try:
        # Get the document record
        document = db.query(Document).filter(Document.id == doc_id).first()
        if not document:
            print(f"[ERROR] Document {doc_id} not found")
            return

        # Update status to processing
        document.processing_status = "processing"
        db.commit()

        print(f"[PROCESSING] Processing: {document.filename}")

        # 1. Extract text
        text, page_count = extract_text(document.file_path, document.file_type)
        if not text:
            document.processing_status = "failed"
            document.summary = "No text could be extracted from this document."
            db.commit()
            return

        print(f"   [OK] Extracted {len(text)} characters, {page_count} pages")

        # 2. Classify document
        category = classify_document(text, document.filename)
        document.doc_category = category
        print(f"   [OK] Classified as: {category}")

        # 3. Chunk text
        chunks = chunk_text(text)
        print(f"   [OK] Created {len(chunks)} chunks")

        # 4. Store in FAISS
        chunk_count = store_chunks_in_faiss(doc_id, chunks, document.filename)
        print(f"   [OK] Stored {chunk_count} chunks in FAISS")

        # 5. Generate a simple summary
        from app.services.llm_service import summarize_text
        summary = summarize_text(text)
        print(f"   [OK] Summarized document")
        
        # 5.5 Extract Entities (NER)
        from app.services.entity_extractor import extract_entities
        entities = extract_entities(text, doc_id)
        print(f"   [OK] Extracted {len(entities)} entities")
        
        # 6. Extract Assets and sync to dashboard DB
        from app.services.llm_service import extract_assets_from_text
        from app.models.models import Asset
        extracted_assets = extract_assets_from_text(text)
        
        for asset_data in extracted_assets:
            # Upsert asset (insert if not exists, update if exists)
            asset_id = asset_data.get("name", "").strip().upper()
            if not asset_id:
                continue
                
            existing = db.query(Asset).filter(Asset.id == asset_id).first()
            if existing:
                existing.health_score = asset_data.get("health_score", existing.health_score)
                existing.status = asset_data.get("status", existing.status)
            else:
                new_asset = Asset(
                    id=asset_id,
                    name=asset_data.get("name"),
                    type=asset_data.get("type"),
                    criticality=asset_data.get("criticality", "Medium"),
                    health_score=asset_data.get("health_score", 85.0),
                    status=asset_data.get("status", "Active"),
                    location="Extracted from Docs"
                )
                db.add(new_asset)
        
        db.commit()
        print(f"   [OK] Extracted and synced {len(extracted_assets)} assets to Dashboard")

        # 6. Update document record
        document.processed = True
        document.processing_status = "ready"
        document.page_count = page_count
        document.chunk_count = chunk_count
        document.summary = summary

        db.commit()
        print(f"   [DONE] Done processing: {document.filename}")

    except Exception as e:
        print(f"[ERROR] Error processing {doc_id}: {str(e)}")
        traceback.print_exc()
        try:
            document = db.query(Document).filter(Document.id == doc_id).first()
            if document:
                document.processing_status = "failed"
                document.summary = f"Processing error: {str(e)}"
                db.commit()
        except Exception:
            pass
    finally:
        db.close()
